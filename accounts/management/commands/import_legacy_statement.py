import math
from decimal import Decimal, InvalidOperation
from typing import Tuple, Dict, List, Any, Optional

import pandas as pd  # opcional: útil para futuras utilidades, no se usa para leer
from dateutil import parser as dateparser
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounts.models import StatementFile, LegacyStatementExcel


# ======== Constantes / mapeos ========
REQUIRED_COLUMNS = [
    "Work Primary Title",
    "Work Writer List",
    "ISWC",
    "Usage Period Start Date",
    "Usage Period End Date",
    "Use Type",
    "Processing Type",
    "DSP Name",
    "Number of Usages",
    "Distributed Amount",
]

COL_MAP = {
    "Work Primary Title": "work_primary_title",
    "Work Writer List": "work_writer_list",
    "ISWC": "iswc",
    "Usage Period Start Date": "usage_period_start",
    "Usage Period End Date": "usage_period_end",
    "Use Type": "use_type",
    "Processing Type": "processing_type",
    "DSP Name": "dsp_name",
    "Number of Usages": "number_of_usages",
    "Distributed Amount": "distributed_amount_usd",
}


# ======== Utilidades de parsing ========
def to_none_if_nan(v):
    # Normaliza strings tipo NaN de Excel
    try:
        if v is None:
            return None
        if isinstance(v, float) and math.isnan(v):
            return None
        s = str(v).strip()
        return None if s.lower() in ("nan", "") else s
    except Exception:
        return v

def parse_date(val):
    val = to_none_if_nan(val)
    if val is None:
        return None
    try:
        if hasattr(val, "to_pydatetime"):
            return val.to_pydatetime().date()
        return dateparser.parse(str(val), dayfirst=True, fuzzy=True).date()
    except Exception:
        return None

def parse_int(val):
    val = to_none_if_nan(val)
    if val is None:
        return None
    try:
        return int(str(val).replace(",", "").strip())
    except Exception:
        try:
            return int(float(val))
        except Exception:
            return None

def parse_decimal(val):
    val = to_none_if_nan(val)
    if val is None:
        return None
    s = str(val).strip()
    for t in ["$", "USD", "usd", "COP", "cop"]:
        s = s.replace(t, "")
    s = s.replace(",", "").replace(" ", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        try:
            return Decimal(str(float(val)))
        except Exception:
            return None


# ======== Detección de encabezados (tolerante) ========
def _norm(s: str) -> str:
    import re
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def detect_header_and_sheet(xlsx_path: str) -> Tuple[str, int]:
    """
    Detecta hoja y fila del encabezado buscando una fila que contenga, de forma
    flexible, 'Work Primary Title' y 'Work Writer List' (normalizados).
    Retorna: (sheet_name, header_row_index)  [header_row_index es 0-based]
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    key_title = _norm("Work Primary Title")
    key_writer = _norm("Work Writer List")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), start=1):
            values = [_norm(str(v)) if v is not None else "" for v in row]
            vals_set = set(values)
            if key_title in vals_set and key_writer in vals_set:
                return sheet, i - 1
    raise CommandError("No pude detectar encabezados (se requieren 'Work Primary Title' y 'Work Writer List').")

def detect_header_row_in_sheet(ws) -> Optional[int]:
    """
    Solo detecta la fila de encabezado dentro de una worksheet ya elegida.
    Devuelve índice 0-based o None si no la encuentra.
    """
    key_title = _norm("Work Primary Title")
    key_writer = _norm("Work Writer List")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), start=1):
        values = [_norm(str(v)) if v is not None else "" for v in row]
        vals = set(values)
        if key_title in vals and key_writer in vals:
            return i - 1
    return None


# ======== Helpers de cabecera / objeto ========
def get_header_positions(header_row: List[Any]) -> Dict[str, int]:
    """
    Devuelve un dict {col_name: index} para las columnas requeridas.
    """
    pos: Dict[str, int] = {}
    for idx, val in enumerate(header_row):
        if val is None:
            continue
        name = str(val).strip()
        if name in REQUIRED_COLUMNS:
            pos[name] = idx
    missing = [c for c in REQUIRED_COLUMNS if c not in pos]
    if missing:
        raise CommandError(f"Faltan columnas en el encabezado: {missing}")
    return pos

def build_obj_from_row(row: List[Any], pos: Dict[str, int], id_file: int) -> LegacyStatementExcel:
    # Extrae y parsea cada campo
    get = lambda col: row[pos[col]] if pos[col] < len(row) else None

    work_primary_title = to_none_if_nan(get("Work Primary Title"))
    work_writer_list = to_none_if_nan(get("Work Writer List"))
    iswc = to_none_if_nan(get("ISWC"))
    usage_period_start = parse_date(get("Usage Period Start Date"))
    usage_period_end = parse_date(get("Usage Period End Date"))
    use_type = to_none_if_nan(get("Use Type"))
    processing_type = to_none_if_nan(get("Processing Type"))
    dsp_name = to_none_if_nan(get("DSP Name"))
    number_of_usages = parse_int(get("Number of Usages"))
    distributed_amount_usd = parse_decimal(get("Distributed Amount"))

    return LegacyStatementExcel(
        id_file=id_file,
        work_writer_list=work_writer_list,
        work_primary_title=work_primary_title,
        iswc=iswc,
        usage_period_start=usage_period_start,
        usage_period_end=usage_period_end,
        use_type=use_type,
        processing_type=processing_type,
        dsp_name=dsp_name,
        number_of_usages=number_of_usages,
        distributed_amount_usd=distributed_amount_usd,
    )


# ======== Comando principal ========
class Command(BaseCommand):
    help = "Importa un Excel legado (XLSX) con las 10 columnas clásicas, leyendo en streaming y cargando por lotes."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Ruta al archivo XLSX (legado)")
        parser.add_argument("--cliente-id", dest="cliente_id", type=int, required=True, help="ID del cliente (FK a clientes)")
        parser.add_argument("--derecho", dest="derecho", type=str, required=True, choices=["Mecanico", "MicroSync"], help="Tipo de derecho")
        parser.add_argument("--anio", dest="anio", type=int, required=True, help="Año del período (ej. 2024)")
        parser.add_argument("--q", dest="q", type=int, required=True, choices=[1, 2, 3, 4], help="Quarter (1..4)")
        parser.add_argument("--nombre-archivo", dest="nombre_archivo", type=str, default=None, help="Nombre lógico a guardar (opcional)")
        parser.add_argument("--dry-run", dest="dry_run", action="store_true", help="No escribe en BD; solo valida y muestra resumen")
        parser.add_argument("--batch-size", dest="batch_size", type=int, default=5000, help="Tamaño de lote para bulk_create")

        # Overrides de UNA sola hoja (opcional)
        parser.add_argument("--sheet-name", dest="sheet_name", type=str, help="Nombre de la hoja (override)")
        parser.add_argument("--header-row", dest="header_row", type=int, help="Fila de encabezado 1-based (override)")

        # Multi-hoja
        parser.add_argument("--sheet-names", dest="sheet_names", type=str, help="Lista de hojas separadas por coma a importar (en orden).")
        parser.add_argument("--all-sheets", dest="all_sheets", action="store_true", help="Importar todas las hojas que contengan las columnas requeridas.")

        # Append a un StatementFile existente
        parser.add_argument("--append-to", dest="append_to_id_file", type=int, help="ID de statement_files al que se añadirán filas (no crea uno nuevo).")

    def handle(self, *args, **opts):
        xlsx_path = opts["xlsx_path"]
        cliente_id = opts["cliente_id"]
        derecho = opts["derecho"]
        anio = opts["anio"]
        periodo_q = opts["q"]
        nombre_archivo = opts.get("nombre_archivo") or xlsx_path.split("\\")[-1].split("/")[-1]
        dry_run = opts.get("dry_run", False)
        batch_size = int(opts.get("batch_size") or 5000)

        sheet_name_override = opts.get("sheet_name")
        header_row_override = opts.get("header_row")
        sheet_names_arg = opts.get("sheet_names")
        all_sheets = bool(opts.get("all_sheets"))
        append_to_id_file = opts.get("append_to_id_file")

        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        # --- 1) Determinar qué hojas procesar ---
        if sheet_names_arg:
            sheets_to_process = [s.strip() for s in sheet_names_arg.split(",") if s.strip()]
        elif all_sheets:
            sheets_to_process = list(wb.sheetnames)
        else:
            # Modo de una sola hoja: usa override si está completo, o autodetección
            if sheet_name_override and header_row_override:
                sheets_to_process = [sheet_name_override]
            else:
                auto_sheet, _ = detect_header_and_sheet(xlsx_path)
                sheets_to_process = [auto_sheet]

        # --- 2) Crear/obtener StatementFile ---
        inserted_total = 0
        if append_to_id_file:
            try:
                sf = StatementFile.objects.get(pk=append_to_id_file)
            except StatementFile.DoesNotExist:
                raise CommandError(f"append_to id_file={append_to_id_file} no existe.")
            prev_count = sf.filas_cargadas or 0  # para calcular correctamente al final
        else:
            sf = StatementFile.objects.create(
                cliente_id=cliente_id,
                derecho=derecho,
                anio=anio,
                periodo_q=periodo_q,
                file_type="XLSX",
                nombre_archivo=nombre_archivo,
                filas_cargadas=0,
            )
            prev_count = 0

        # --- 3) Procesar cada hoja ---
        for sheet_name in sheets_to_process:
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f"Ignorando '{sheet_name}': no existe en el workbook."))
                continue

            ws = wb[sheet_name]

            # Elegir header_idx: override solo aplica si coincide la hoja
            if sheet_name_override == sheet_name and header_row_override:
                header_idx = int(header_row_override) - 1  # 0-based
            else:
                header_idx = detect_header_row_in_sheet(ws)

            if header_idx is None:
                # Si la hoja no contiene las columnas requeridas, la saltamos
                self.stdout.write(self.style.WARNING(
                    f"Saltando hoja '{sheet_name}': no encontré encabezado con las columnas requeridas."
                ))
                continue

            # Leer cabecera y posiciones
            header_row_cells = next(ws.iter_rows(min_row=header_idx + 1, max_row=header_idx + 1, values_only=True))
            pos = get_header_positions(list(header_row_cells))

            if dry_run:
                # Contar filas y mostrar 5 ejemplos
                total = 0
                samples: List[Dict[str, Any]] = []
                for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
                    if row is None:
                        continue
                    total += 1
                    if len(samples) < 5:
                        obj = build_obj_from_row(list(row), pos, id_file=-1)
                        samples.append({
                            "Work Primary Title": obj.work_primary_title,
                            "Work Writer List": obj.work_writer_list,
                            "ISWC": obj.iswc,
                            "Usage Period Start Date": obj.usage_period_start,
                            "Usage Period End Date": obj.usage_period_end,
                            "Use Type": obj.use_type,
                            "Processing Type": obj.processing_type,
                            "DSP Name": obj.dsp_name,
                            "Number of Usages": obj.number_of_usages,
                            "Distributed Amount": obj.distributed_amount_usd,
                        })
                self.stdout.write(self.style.NOTICE(
                    f"[{sheet_name}] Encabezado en fila: {header_idx+1} | Filas: {total}"
                ))
                for i, row in enumerate(samples, start=1):
                    self.stdout.write(f"  {i}. {row}")
                # sigue con la próxima hoja
                continue

            # Carga real por lotes
            batch: List[LegacyStatementExcel] = []
            for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
                if row is None:
                    continue
                batch.append(build_obj_from_row(list(row), pos, id_file=sf.id_file))
                if len(batch) >= batch_size:
                    with transaction.atomic():
                        LegacyStatementExcel.objects.bulk_create(batch, batch_size=batch_size)
                    inserted_total += len(batch)
                    batch.clear()

            if batch:
                with transaction.atomic():
                    LegacyStatementExcel.objects.bulk_create(batch, batch_size=batch_size)
                inserted_total += len(batch)
                batch.clear()

        # --- 4) Finalizar ---
        if dry_run:
            self.stdout.write(self.style.SUCCESS("Dry-run finalizado."))
            return

        # Ajuste final CORRECTO del contador
        if append_to_id_file:
            sf.filas_cargadas = prev_count + inserted_total
        else:
            sf.filas_cargadas = inserted_total
        sf.save(update_fields=["filas_cargadas"])

        self.stdout.write(self.style.SUCCESS(
            f"Importación completada: id_file={sf.id_file}, filas_cargadas={sf.filas_cargadas}"
        ))
