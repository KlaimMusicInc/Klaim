# accounts/management/commands/import_microsync_statement.py
import math
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

from dateutil import parser as dateparser
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounts.models import (MicroSyncMarketShare, MicroSyncStatement,
                             StatementFile)

# ============== Columnas requeridas =================
DETAIL_REQUIRED = [
    "Asset Title",
    "Asset Type",
    "Track Code",
    "Artist",
    "Type",
    "Country",
    "AD Total Views",
    "Amount Payable",
]

MARKET_REQUIRED = [
    "Description",
    "Amount Payable",
]


# ============== Utilidades de parsing =================
def to_none_if_nan(v):
    try:
        if v is None:
            return None
        if isinstance(v, float) and math.isnan(v):
            return None
        s = str(v).strip()
        return None if s.lower() in ("nan", "") else s
    except Exception:
        return v


def parse_date(val):
    val = to_none_if_nan(val)
    if val is None:
        return None
    try:
        if hasattr(val, "to_pydatetime"):
            return val.to_pydatetime().date()
        return dateparser.parse(str(val), dayfirst=True, fuzzy=True).date()
    except Exception:
        return None


def parse_int(val):
    val = to_none_if_nan(val)
    if val is None:
        return None
    try:
        return int(str(val).replace(",", "").strip())
    except Exception:
        try:
            return int(float(val))
        except Exception:
            return None


def parse_bigint(val):
    v = parse_int(val)
    return v


def parse_decimal(val):
    val = to_none_if_nan(val)
    if val is None:
        return None
    s = str(val).strip()
    for t in ["$", "USD", "usd", "COP", "cop"]:
        s = s.replace(t, "")
    s = s.replace(",", "").replace(" ", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        try:
            return Decimal(str(float(val)))
        except Exception:
            return None


# ============== Normalización y detección de encabezado =================
def _norm(s: str) -> str:
    import re

    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _row_has_headers(row_vals: List[Any], required: List[str]) -> bool:
    values = [_norm(str(v)) if v is not None else "" for v in row_vals]
    vals_set = set(values)
    return all(_norm(col) in vals_set for col in required)


def detect_header_row(ws, required: List[str]) -> Optional[int]:
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=200, values_only=True), start=1
    ):
        if _row_has_headers(list(row), required):
            return i - 1  # 0-based
    return None


def find_sheet_and_header(
    wb, required: List[str]
) -> Tuple[Optional[str], Optional[int]]:
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        idx = detect_header_row(ws, required)
        if idx is not None:
            return sheet, idx
    return None, None


def get_header_positions(header_row: List[Any], required: List[str]) -> Dict[str, int]:
    pos: Dict[str, int] = {}
    for idx, val in enumerate(header_row):
        if val is None:
            continue
        name = str(val).strip()
        if name in required:
            pos[name] = idx
    missing = [c for c in required if c not in pos]
    if missing:
        raise CommandError(f"Faltan columnas en el encabezado: {missing}")
    return pos


# ============== Builders =================
def build_ms_from_row(
    row: List[Any], pos: Dict[str, int], id_file: int
) -> MicroSyncStatement:
    get = lambda col: row[pos[col]] if pos[col] < len(row) else None
    return MicroSyncStatement(
        id_file=id_file,
        asset_title=to_none_if_nan(get("Asset Title")),
        asset_type=to_none_if_nan(get("Asset Type")),
        track_code=to_none_if_nan(get("Track Code")),
        artist=to_none_if_nan(get("Artist")),
        type=to_none_if_nan(get("Type")),
        country=to_none_if_nan(get("Country")),
        ad_total_views=parse_bigint(get("AD Total Views")),
        amount_payable_usd=parse_decimal(get("Amount Payable")),
    )


def build_share_from_row(
    row: List[Any], pos: Dict[str, int], id_file: int
) -> MicroSyncMarketShare:
    get = lambda col: row[pos[col]] if pos[col] < len(row) else None
    return MicroSyncMarketShare(
        id_file=id_file,
        description=to_none_if_nan(get("Description")),
        amount_payable_usd=parse_decimal(get("Amount Payable")),
    )


# ============== Comando =================
class Command(BaseCommand):
    help = "Importa XLSX de MicroSync: hoja de detalle (Asset...) y hoja de cuota de mercado (Description/Amount Payable). Soporta anual (Q=0)."

    def add_arguments(self, parser):
        parser.add_argument(
            "xlsx_path", type=str, help="Ruta al archivo XLSX (MicroSync)"
        )
        parser.add_argument("--cliente-id", dest="cliente_id", type=int, required=True)
        parser.add_argument("--anio", dest="anio", type=int, required=True)

        # Q es OPCIONAL. Si no se da, usa --annual para periodo_q=0
        parser.add_argument(
            "--q",
            dest="q",
            type=int,
            choices=[1, 2, 3, 4],
            help="Quarter (1..4). Si no se indica, usa --annual.",
        )
        parser.add_argument(
            "--annual",
            dest="annual",
            action="store_true",
            help="Archivo anual (periodo_q=0).",
        )

        parser.add_argument(
            "--nombre-archivo", dest="nombre_archivo", type=str, default=None
        )
        parser.add_argument(
            "--dry-run",
            dest="dry_run",
            action="store_true",
            help="No escribe en BD; solo valida y muestra resumen",
        )
        parser.add_argument("--batch-size", dest="batch_size", type=int, default=5000)

        # Overrides de hojas / encabezados
        parser.add_argument(
            "--detalle-sheet",
            dest="detalle_sheet",
            type=str,
            help="Nombre hoja DETALLE",
        )
        parser.add_argument(
            "--detalle-header-row",
            dest="detalle_header_row",
            type=int,
            help="Fila encabezado DETALLE (1-based)",
        )
        parser.add_argument(
            "--cuota-sheet", dest="cuota_sheet", type=str, help="Nombre hoja CUOTA"
        )
        parser.add_argument(
            "--cuota-header-row",
            dest="cuota_header_row",
            type=int,
            help="Fila encabezado CUOTA (1-based)",
        )

    def handle(self, *args, **opts):
        xlsx_path = opts["xlsx_path"]
        cliente_id = opts["cliente_id"]
        anio = opts["anio"]
        annual = bool(opts.get("annual"))
        q_opt = opts.get("q")

        # Validación de combinaciones
        if annual and q_opt:
            raise CommandError("No combines --annual con --q; usa uno u otro.")
        if not annual and not q_opt:
            raise CommandError(
                "Debes indicar --q 1..4 o usar --annual para archivo anual."
            )

        periodo_q = 0 if annual else int(q_opt)

        nombre_archivo = (
            opts.get("nombre_archivo") or xlsx_path.split("\\")[-1].split("/")[-1]
        )
        dry_run = bool(opts.get("dry_run"))
        batch_size = int(opts.get("batch_size") or 5000)

        detalle_sheet = opts.get("detalle_sheet")
        detalle_header_row = opts.get("detalle_header_row")
        cuota_sheet = opts.get("cuota_sheet")
        cuota_header_row = opts.get("cuota_header_row")

        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        # --- Detectar hoja/encabezado DETALLE ---
        if detalle_sheet and detalle_header_row:
            if detalle_sheet not in wb.sheetnames:
                raise CommandError(f"Hoja detalle '{detalle_sheet}' no existe.")
            ws_det = wb[detalle_sheet]
            det_header_idx = int(detalle_header_row) - 1
        else:
            sheet_auto, det_header_idx = find_sheet_and_header(wb, DETAIL_REQUIRED)
            if not sheet_auto:
                raise CommandError(
                    "No se encontró hoja con columnas de DETALLE requeridas."
                )
            ws_det = wb[sheet_auto]

        # --- Detectar hoja/encabezado CUOTA ---
        if cuota_sheet and cuota_header_row:
            if cuota_sheet not in wb.sheetnames:
                raise CommandError(f"Hoja cuota '{cuota_sheet}' no existe.")
            ws_cuota = wb[cuota_sheet]
            cuota_header_idx = int(cuota_header_row) - 1
        else:
            sheet_auto, cuota_header_idx = find_sheet_and_header(wb, MARKET_REQUIRED)
            if not sheet_auto:
                # La cuota puede no venir; toleramos pero avisamos
                ws_cuota = None
                cuota_header_idx = None
            else:
                ws_cuota = wb[sheet_auto]

        # --- Leer cabeceras y posiciones ---
        det_header_cells = next(
            ws_det.iter_rows(
                min_row=det_header_idx + 1, max_row=det_header_idx + 1, values_only=True
            )
        )
        det_pos = get_header_positions(list(det_header_cells), DETAIL_REQUIRED)

        cuota_pos = None
        if ws_cuota is not None and cuota_header_idx is not None:
            cuota_header_cells = next(
                ws_cuota.iter_rows(
                    min_row=cuota_header_idx + 1,
                    max_row=cuota_header_idx + 1,
                    values_only=True,
                )
            )
            cuota_pos = get_header_positions(list(cuota_header_cells), MARKET_REQUIRED)

        # --- Dry-run: contar y samplear ---
        if dry_run:
            total_det = 0
            samples_det: List[Dict[str, Any]] = []
            for row in ws_det.iter_rows(min_row=det_header_idx + 2, values_only=True):
                if row is None:
                    continue
                total_det += 1
                if len(samples_det) < 5:
                    o = build_ms_from_row(list(row), det_pos, id_file=-1)
                    samples_det.append(
                        {
                            "Asset Title": o.asset_title,
                            "Asset Type": o.asset_type,
                            "Track Code": o.track_code,
                            "Artist": o.artist,
                            "Type": o.type,
                            "Country": o.country,
                            "AD Total Views": o.ad_total_views,
                            "Amount Payable": o.amount_payable_usd,
                        }
                    )

            label_periodo = f"Q{periodo_q}" if periodo_q else "ANUAL (Q=0)"
            self.stdout.write(
                self.style.NOTICE(
                    f"[DETALLE] Hoja '{ws_det.title}' encabezado fila {det_header_idx+1} | Filas: {total_det} | Período: {label_periodo}"
                )
            )
            for i, row in enumerate(samples_det, start=1):
                self.stdout.write(f"  {i}. {row}")

            if ws_cuota and cuota_pos:
                total_cuota = 0
                samples_cuota: List[Dict[str, Any]] = []
                for row in ws_cuota.iter_rows(
                    min_row=cuota_header_idx + 2, values_only=True
                ):
                    if row is None:
                        continue
                    total_cuota += 1
                    if len(samples_cuota) < 5:
                        s = build_share_from_row(list(row), cuota_pos, id_file=-1)
                        samples_cuota.append(
                            {
                                "Description": s.description,
                                "Amount Payable": s.amount_payable_usd,
                            }
                        )
                self.stdout.write(
                    self.style.NOTICE(
                        f"[CUOTA] Hoja '{ws_cuota.title}' encabezado fila {cuota_header_idx+1} | Filas: {total_cuota}"
                    )
                )
                for i, row in enumerate(samples_cuota, start=1):
                    self.stdout.write(f"  {i}. {row}")

            self.stdout.write(self.style.SUCCESS("Dry-run finalizado."))
            return

        # --- Crear StatementFile ---
        sf = StatementFile.objects.create(
            cliente_id=cliente_id,
            derecho="MicroSync",
            anio=anio,
            periodo_q=periodo_q,  # 0 si anual
            file_type="XLSX",
            nombre_archivo=nombre_archivo,
            filas_cargadas=0,
        )

        # --- Insertar DETALLE por lotes ---
        inserted_det = 0
        batch: List[MicroSyncStatement] = []
        for row in ws_det.iter_rows(min_row=det_header_idx + 2, values_only=True):
            if row is None:
                continue
            batch.append(build_ms_from_row(list(row), det_pos, id_file=sf.id_file))
            if len(batch) >= batch_size:
                with transaction.atomic():
                    MicroSyncStatement.objects.bulk_create(batch, batch_size=batch_size)
                inserted_det += len(batch)
                batch.clear()
        if batch:
            with transaction.atomic():
                MicroSyncStatement.objects.bulk_create(batch, batch_size=batch_size)
            inserted_det += len(batch)
            batch.clear()

        # --- Insertar CUOTA (si existe) por lotes ---
        inserted_cuota = 0
        if ws_cuota and cuota_pos:
            batch_share: List[MicroSyncMarketShare] = []
            for row in ws_cuota.iter_rows(
                min_row=cuota_header_idx + 2, values_only=True
            ):
                if row is None:
                    continue
                batch_share.append(
                    build_share_from_row(list(row), cuota_pos, id_file=sf.id_file)
                )
                if len(batch_share) >= batch_size:
                    with transaction.atomic():
                        MicroSyncMarketShare.objects.bulk_create(
                            batch_share, batch_size=batch_size
                        )
                    inserted_cuota += len(batch_share)
                    batch_share.clear()
            if batch_share:
                with transaction.atomic():
                    MicroSyncMarketShare.objects.bulk_create(
                        batch_share, batch_size=batch_size
                    )
                inserted_cuota += len(batch_share)
                batch_share.clear()

        # --- Actualizar contador y log ---
        sf.filas_cargadas = inserted_det + inserted_cuota  # total en ambas tablas
        sf.save(update_fields=["filas_cargadas"])

        self.stdout.write(
            self.style.SUCCESS(
                f"Importación completada: id_file={sf.id_file}, detalle={inserted_det}, cuota={inserted_cuota}, total={sf.filas_cargadas}"
            )
        )
